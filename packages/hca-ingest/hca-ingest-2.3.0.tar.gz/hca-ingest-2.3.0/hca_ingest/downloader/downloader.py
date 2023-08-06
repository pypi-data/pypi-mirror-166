from typing import List

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from hca_ingest.downloader.flattener import Flattener
from hca_ingest.importer.spreadsheet.ingest_workbook import SCHEMAS_WORKSHEET
from hca_ingest.importer.spreadsheet.ingest_worksheet import START_DATA_ROW

HEADER_ROW_NO = 4


class XlsDownloader:
    def __init__(self):
        self.flattener = Flattener()

    def convert_json(self, metadata_list: List[dict]):
        return self.flattener.flatten(metadata_list)

    def create_workbook(self, input_json: dict) -> Workbook:
        workbook = Workbook()
        workbook.remove(workbook.active)

        for ws_title, ws_elements in input_json.items():
            if ws_title == 'Project':
                worksheet: Worksheet = workbook.create_sheet(title=ws_title, index=0)
            elif ws_title == SCHEMAS_WORKSHEET:
                continue
            else:
                worksheet: Worksheet = workbook.create_sheet(title=ws_title)

            self.add_worksheet_content(worksheet, ws_elements)

        self.generate_schemas_worksheet(input_json, workbook)

        return workbook

    def generate_schemas_worksheet(self, input_json, workbook):
        schemas = input_json.get(SCHEMAS_WORKSHEET)
        if not schemas:
            raise ValueError('The schema urls are missing')
        schemas_worksheet = workbook.create_sheet(SCHEMAS_WORKSHEET)
        schemas_worksheet.cell(row=1, column=1, value=SCHEMAS_WORKSHEET)
        for row_num, schema in enumerate(schemas, start=2):
            schemas_worksheet.cell(row=row_num, column=1, value=schema)

    def add_worksheet_content(self, worksheet, ws_elements: dict):
        headers = ws_elements.get('headers')
        self.__add_header_row(worksheet, headers)
        all_values = ws_elements.get('values')

        for row_number, row_values in enumerate(all_values, start=START_DATA_ROW):
            self.__add_row_content(worksheet, headers, row_number, row_values)

    @staticmethod
    def __add_header_row(worksheet, headers: list):
        for col, header in enumerate(headers, start=1):
            worksheet.cell(row=HEADER_ROW_NO, column=col, value=header)

    @staticmethod
    def __add_row_content(worksheet, headers: list, row_number: int, values: dict):
        for header, value in values.items():
            index = headers.index(header)
            worksheet.cell(row=row_number, column=index + 1, value=value)
