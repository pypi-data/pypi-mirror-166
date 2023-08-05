import atexit
import fcntl
import glob
import json
import logging
import logging.handlers
import logging.handlers
import os
import re
import signal
import sys
import time
import traceback
from argparse import Namespace
from collections import OrderedDict
from json import JSONDecodeError
from multiprocessing import Lock
from timeit import default_timer as timer

import pandas as pd
from openpyxl.utils import get_column_letter
from watchdog.events import FileSystemEventHandler
from watchdog.observers import Observer


def signal_handler(signal, frame):
    print('You pressed Ctrl+C!')
    os._exit(0)


signal.signal(signal.SIGINT, signal_handler)


class runOnlyOnce():
    """
    usage:
        runOnlyOnce()
    """

    def __init__(self, path=None):
        self.path = path if path is not None else os.path.basename(__file__).replace(".py", ".pid")
        self.pidfile = open(self.path, "a+")
        try:
            fcntl.flock(self.pidfile.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
        except IOError as e:
            raise SystemExit("Already running according to " + self.path)

        atexit.register(self.log_exit)
        self.pidfile.seek(0)
        self.pidfile.truncate()
        self.pidfile.write(str(os.getpid()))
        self.pidfile.flush()
        self.pidfile.seek(0)

    def log_exit(self):
        try:
            self.pidfile.close()
        except IOError as err:
            if err.errno != 9:
                raise
        os.remove(self.path)
        print("Main process terminated")


class ParallelTimedRotatingFileHandler(logging.handlers.TimedRotatingFileHandler):
    def __init__(self, filename=None, when='h', interval=1, backupCount=0, encoding=None, delay=False, utc=False, atTime=None, postfix=".log"):
        self.when = when.upper()
        self.interval = interval
        self.backupCount = backupCount
        self.utc = utc
        self.postfix = postfix
        self.atTime = atTime

        if self.when == 'S':
            self.interval = 1  # one second
            self.suffix = "%Y-%m-%d_%H-%M-%S" if filename is None else filename
            self.extMatch = r"^\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2}$"
        elif self.when == 'M':
            self.interval = 60  # one minute
            self.suffix = "%Y-%m-%d_%H-%M" if filename is None else filename
            self.extMatch = r"^\d{4}-\d{2}-\d{2}_\d{2}-\d{2}$"
        elif self.when == 'H':
            self.interval = 60 * 60  # one hour
            self.suffix = "%Y-%m-%d_%H" if filename is None else filename
            self.extMatch = r"^\d{4}-\d{2}-\d{2}_\d{2}$"
        elif self.when == 'D' or self.when == 'MIDNIGHT':
            self.interval = 60 * 60 * 24  # one day
            self.suffix = "%Y-%m-%d" if filename is None else filename
            self.extMatch = r"^\d{4}-\d{2}-\d{2}$"
        elif self.when.startswith('W'):
            self.interval = 60 * 60 * 24 * 7  # one week
            if len(self.when) != 2:
                raise ValueError("You must specify a day for weekly rollover from 0 to 6 (0 is Monday): %s" % self.when)
            if self.when[1] < '0' or self.when[1] > '6':
                raise ValueError("Invalid day specified for weekly rollover: %s" % self.when)
            self.dayOfWeek = int(self.when[1])
            self.suffix = "%Y-%m-%d" if filename is None else filename
            self.extMatch = r"^\d{4}-\d{2}-\d{2}$"
        else:
            raise ValueError("Invalid rollover interval specified: %s" % self.when)

        currenttime = int(time.time())
        logging.handlers.BaseRotatingHandler.__init__(self, self.calculateFileName(currenttime), 'a', encoding, delay)

        self.extMatch = re.compile(self.extMatch)
        self.interval = self.interval * interval  # multiply by units requested

        self.rolloverAt = self.computeRollover(currenttime)

    def calculateFileName(self, currenttime):
        if self.utc:
            timeTuple = time.gmtime(currenttime)
        else:
            timeTuple = time.localtime(currenttime)

        return time.strftime(self.suffix, timeTuple) + self.postfix

    def getFilesToDelete(self, newFileName):
        dirName, fName = os.path.split(self.suffix)
        dName, newFileName = os.path.split(newFileName)

        fileNames = os.listdir(dirName)
        result = []
        prefix = fName + "."
        postfix = self.postfix
        prelen = len(prefix)
        postlen = len(postfix)
        for fileName in fileNames:
            if fileName[:prelen] == prefix and fileName[-postlen:] == postfix and len(fileName) - postlen > prelen and fileName != newFileName:
                suffix = fileName[prelen:len(fileName) - postlen]
                if self.extMatch.match(suffix):
                    result.append(os.path.join(dirName, fileName))
        result.sort()
        if len(result) < self.backupCount:
            result = []
        else:
            result = result[:len(result) - self.backupCount]
        return result

    def doRollover(self):
        if self.stream:
            self.stream.close()
            self.stream = None

        currentTime = self.rolloverAt
        newFileName = self.calculateFileName(currentTime)
        newBaseFileName = os.path.abspath(newFileName)
        self.baseFilename = newBaseFileName
        self.mode = 'a'
        self.stream = self._open()

        if self.backupCount > 0:
            for s in self.getFilesToDelete(newFileName):
                try:
                    print(s)
                    # os.remove(s)
                except:
                    pass

        newRolloverAt = self.computeRollover(currentTime)
        while newRolloverAt <= currentTime:
            newRolloverAt = newRolloverAt + self.interval

        # If DST changes and midnight or weekly rollover, adjust for this.
        if (self.when == 'MIDNIGHT' or self.when.startswith('W')) and not self.utc:
            dstNow = time.localtime(currentTime)[-1]
            dstAtRollover = time.localtime(newRolloverAt)[-1]
            if dstNow != dstAtRollover:
                if not dstNow:  # DST kicks in before next rollover, so we need to deduct an hour
                    newRolloverAt = newRolloverAt - 3600
                else:  # DST bows out before next rollover, so we need to add an hour
                    newRolloverAt = newRolloverAt + 3600
        self.rolloverAt = newRolloverAt


class Singleton(type):
    _instances = {}

    def __call__(cls, *args, **kwargs):
        if cls not in cls._instances:
            cls._instances[cls] = super(Singleton, cls).__call__(*args, **kwargs)
        return cls._instances[cls]


class SetupLogger(metaclass=Singleton):
    def __init__(self, path="log", error_path="log", level=logging.DEBUG):
        if not os.path.isabs(path):
            path = os.path.join(os.path.dirname(os.path.realpath(__file__)), path)
        if not os.path.isabs(error_path):
            error_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), error_path)

        os.makedirs(path, exist_ok=1)
        os.makedirs(error_path, exist_ok=1)

        logger = logging.getLogger(sys.argv[0])
        logger.setLevel(level)

        formatter = logging.Formatter("%(asctime)s %(filename)10s:%(lineno)3d [%(levelname)8s] %(message)s", datefmt='%Y-%m-%d %H:%M:%S')
        logHandler = ParallelTimedRotatingFileHandler(filename=path + os.sep + 'NORMAL_WEEK_%02V', when='W0', backupCount=0, encoding='utf8')
        logHandler.setLevel(logging.INFO)
        logHandler.setFormatter(formatter)

        errorLogHandler = ParallelTimedRotatingFileHandler(error_path + os.sep + 'ERROR_WEEK_%02V', when='W0', backupCount=0, encoding='utf8')
        errorLogHandler.setLevel(logging.ERROR)
        errorLogHandler.setFormatter(formatter)

        streamHandler = logging.StreamHandler()
        streamHandler.setLevel(logging.INFO)
        streamHandler.setFormatter(formatter)

        logger.addHandler(logHandler)
        logger.addHandler(errorLogHandler)
        logger.addHandler(streamHandler)

        self.info = logger.info
        self.warning = logger.warning
        self.warn = logger.warning
        self.error = logger.error
        self.debug = logger.debug
        self.critical = logger.critical
        self.info("setting up logger")


log = SetupLogger(level=logging.DEBUG)


def debugger_is_active():
    return hasattr(sys, 'gettrace') and sys.gettrace() is not None


def check_file_writable(fnm):
    if os.path.exists(fnm):
        # path exists
        if os.path.isfile(fnm):  # is it a file or a dir?
            # also works when file is a link and the target is writable
            return os.access(fnm, os.W_OK)
        else:
            sys.exit("ERROR: Can not open file %s because it's a folder" % fnm)
            return False  # path is a dir, so cannot write as a file
    # target does not exist, check perms on parent dir
    pdir = os.path.dirname(fnm)
    if not pdir: pdir = '.'
    # target is creatable if parent dir is writable
    if not os.access(pdir, os.W_OK):
        sys.exit("Can not open file %s for writing." % fnm)
    return os.access(pdir, os.W_OK)


def is_float(element):
    try:
        float(element)
        return True
    except ValueError:
        return False


class Config(OrderedDict, metaclass=Singleton):
    __data = OrderedDict()

    def __init__(self, autoSave=True, default={}, file="config.json", *args, **kwargs):
        super(Config, self).__init__(*args, **kwargs)
        self.__data.__autoSave = autoSave
        self.__data.__file = os.path.join(os.path.dirname(sys.argv[0]), file)
        self.__data.__default = default
        if not os.path.isfile(self.__data.__file):
            self.create_default()
        self.loadConfig()

    def __setattr__(self, key, value):
        # print ("Setting 1",key,value)
        if key in self:
            del self[key]
        OrderedDict.__setitem__(self, key, value)
        if self.__data.__autoSave:
            self.saveConfig()

    def __setitem__(self, key, value):
        # print ("Setting 2",key,value)
        if key in self:
            del self[key]
        OrderedDict.__setitem__(self, key, value)
        if self.__data.__autoSave:
            self.saveConfig()

    def __getattr__(self, item):
        try:
            return self.__getitem__(item)
        except KeyError:
            raise AttributeError(item)

    def create_default(self):
        with open(self.__data.__file, 'w') as outfile:
            json.dump(self.__data.__default, outfile, indent=2)

    def delete(self):
        if os.path.isfile(self.__data.__file):
            os.unlink(self.__data.__file)
            for key in self:
                OrderedDict.__delitem__(self, key)

    def loadConfig(self):
        try:
            with open(self.__data.__file) as json_data:
                self.update(json.load(json_data, object_pairs_hook=OrderedDict))
        except JSONDecodeError:
            print("Invalid JSON, restoring default")
            self.create_default()

    def saveConfig(self):
        print("saving config %s" % self.__data.__file)
        with open(self.__data.__file, 'w') as outfile:
            json.dump(self, outfile, indent=2)


class benchmark(object):
    def __init__(self, msg, fmt="%0.3g"):
        self.msg = msg
        self.fmt = fmt

    def __enter__(self):
        self.start = timer()
        return self

    def __exit__(self, *args):
        t = timer() - self.start
        print(("%s : " + self.fmt + " seconds") % (self.msg, t))
        self.time = t


def adjust_column_width_from_col(ws):
    for i, col in enumerate(ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=1)):
        width = 0
        for cell in col:
            value = cell.value
            if value is not None:
                if isinstance(value, str) is False:
                    value = str(value)
                width = max(width, len(value))

        col_name = get_column_letter(1 + i)
        ws.column_dimensions[col_name].width = width + 2


def read_excel(file="logic.xlsx"):
    # load excel, skip empty rows
    if not os.path.isfile(file):
        sys.exit("Please use %s to specify the logic and settings." % file)

    logic = pd.read_excel(file, sheet_name="logic").dropna(axis=0, how='all').to_dict("records")
    settings = pd.read_excel(file, sheet_name="settings").dropna(axis=0, how='all').to_dict("records")

    cfg = {c['setting'].strip(): c['value'] for c in settings}

    docTypes = list(set([l['Doc type'].strip() for l in logic]))
    folders = [l['Folder'].strip() for l in logic]
    formats = [l['Document number format'] for l in logic]

    # check for rows with empty document type fields
    if len([l for l in docTypes if pd.isna(l)]):
        sys.exit("Document type with empth field")

    # check for rows with empty folders fields
    if len([l for l in folders if pd.isna(l)]):
        sys.exit("Missing Folder entry")

    if "rejects" in docTypes:
        docTypes.remove("rejects")

    return logic, folders, docTypes, formats, cfg


def write_excel(rows, excelFile, sheet='sheet1'):
    df = pd.DataFrame(rows)

    if not isinstance(rows, list):
        print("WARN: save_to_excel requires a list of dictionaries")

    if not os.path.isfile(excelFile):
        df.to_excel(excelFile, sheet_name=sheet, index=False)

        # time.sleep(1)
        # with pd.ExcelWriter(excelFile) as writer:
        #     for column in df:
        #         column_length = max(df[column].astype(str).map(len).max(), len(column))
        #         col_idx = df.columns.get_loc(column)
        #         writer.sheets[sheet].set_column(col_idx, col_idx, column_length)
        # log.info("created  excel %s: \n%s"%(excelFile, df.to_string()))
    else:
        with pd.ExcelWriter(excelFile, engine="openpyxl", mode='a', if_sheet_exists='overlay') as writer:
            df.to_excel(writer, sheet_name=sheet, startrow=writer.sheets[sheet].max_row, header=None, index=False)
            ws = writer.sheets[sheet]
            # for i in range(1, ws.max_column + 1):
            #     ws.column_dimensions[get_column_letter(i)].bestFit = True
            #     ws.column_dimensions[get_column_letter(i)].auto_size = True
            adjust_column_width_from_col(ws)
            # adjust_width(ws)


class WatchDog:
    def __init__(self, inPath, outPath, cb, fileType="*.pdf", moveFiles=True, recursive=False):
        self.cb = cb
        self.rec = recursive
        self.fileType = fileType
        self.moveFiles = moveFiles
        self.lastChange = None
        self.inPath = inPath
        self.outPath = outPath
        os.makedirs(self.outPath, exist_ok=1)
        self.waitPeriod = 20
        self.observer = Observer()
        self.filesToDo = []
        self.lock = Lock()
        self.running = 1

    def start(self):
        event_handler = FileSystemEventHandler()
        event_handler.on_created = self.on_created

        log.info("== Start monitoring folder %s" % self.inPath)
        self.observer.schedule(event_handler, self.inPath, recursive=self.rec)
        self.observer.start()

        ## now that we are monitoring, let's process pre-existing files
        if self.rec:
            files = [f for f in glob.glob('%s/**/*' % self.inPath, recursive=self.rec) if os.path.splitext(f)[1].lower() == '.pdf']
        else:
            files = [f for f in glob.glob('%s/*' % self.inPath) if os.path.splitext(f)[1].lower() == '.pdf']

        if files != []:
            log.info("== Watchdog running for %d pre-existing files" % len(files))
            for file in files:
                self.action(file)

        try:
            while self.running:
                if self.lastChange is not None:
                    if time.time() - self.lastChange > self.waitPeriod:
                        with self.lock:
                            self.doingFiles = self.filesToDo
                            self.filesToDo = []

                        log.info("Processing new %d files" % len(self.doingFiles))
                        for file in self.doingFiles:
                            log.info(" > %s" % file)
                            self.action(file)
                        self.lastChange = None
                time.sleep(1)
        except KeyboardInterrupt as e:
            log.info("== Ctrl-C pressed, aborting")
            self.observer.unschedule_all()
            self.observer.stop()
        except:
            e = sys.exc_info()[0]
            log.error("== WatchDog exception during operations:%s" % str(e))
            self.observer.unschedule_all()
            self.observer.stop()

        self.observer.join()
        log.info("== WatchDog stopped")

    def stop(self):
        self.running = 0
        self.observer.stop()

    def on_created(self, event):
        if event.is_directory or not event.src_path.lower().endswith(".pdf"):
            return None

        log.info("Detected new file [%s] - waiting %d seconds" % (event.src_path, self.waitPeriod))
        with self.lock:
            self.filesToDo.append(event.src_path)
        self.lastChange = time.time()

    def action(self, files):
        try:
            res = self.cb(files)
        except KeyboardInterrupt as e:
            log.error("Error caught during process: %s" % str(e))
            log.error(traceback.format_exc())
        if self.moveFiles:
            for file in files:
                if os.access(file, os.W_OK):
                    dst = self.outPath + os.sep + os.path.basename(file)
                    os.rename(file, dst)
                else:
                    log.warning("! Can not move completed files: no write access to file %s" % file)


def convert_images_to_pdf(images, pdf1_filename):
    im1 = images[0]
    if len(images) > 1:
        im_list = images[1:]
    else:
        im_list = []
    im1.save(pdf1_filename, "PDF", resolution=100.0, save_all=True, append_images=im_list)
