__version__ = "0.2.0"
#!/usr/bin/env python

import argparse
from datetime import date, time, datetime
import logging
import sys
from importlib_resources import files, as_file
from blackduck.HubRestApi import HubInstance


"""
Created on APR 14, 2022

@author: Dinesh Ravi

BOM diff for the report

Drawback:
Snippet matches and source code match difference is not visible as it syncsup across all versions of the project

required: template_aosd_new.xlsx

py get_bom_components_diff.py ASTERIX2 AED2_ANDROID_S_2_2022-04-05_02-20 AED2_ANDROID_S_2_2021-12-16_02-20  -l 200


"""


def main():
    parser = argparse.ArgumentParser(
        "Retreive BOM component info for the given project and version"
    )
    parser.add_argument("project_name")
    parser.add_argument("version")
    parser.add_argument("oldversion")
    parser.add_argument("-o", help="Output directory", required=True)
    group = parser.add_mutually_exclusive_group()
    group.add_argument(
        "-l",
        "--limit",
        default=500,
        help="Set limit on number of components to retrieve",
    )

    group.add_argument("-u", "--unreviewed", action="store_true")
    group.add_argument("-r", "--reviewed", action="store_true")
    parser.add_argument(
        "-v",
        "--vulnerabilities",
        action="store_true",
        help="Get the vulnerability info for each of the components",
    )
    parser.add_argument(
        "-c",
        "--custom_fields",
        action="store_true",
        help="Get the custom field info for each of the components",
    )

    args = parser.parse_args()

    logging.basicConfig(
        format="%(asctime)s:%(levelname)s:%(message)s",
        stream=sys.stderr,
        level=logging.DEBUG,
    )
    logging.getLogger("requests").setLevel(logging.WARNING)
    logging.getLogger("urllib3").setLevel(logging.WARNING)

    hub = HubInstance()

    project = hub.get_project_by_name(args.project_name)

    print(project.values())

    version = hub.get_version_by_name(project, args.version)
    # print(dir(version))
    oldversion = hub.get_version_by_name(project, args.oldversion)

    components_url = hub.get_link(version, "components")
    oldcomponents_url = hub.get_link(oldversion, "components")

    print(args.limit)
    print(components_url)
    print(oldcomponents_url)
    components_url = f"{components_url}?limit={args.limit}"
    oldcomponents_url = f"{oldcomponents_url}?limit={args.limit}"

    custom_headers = {
        "Accept": "application/vnd.blackducksoftware.bill-of-materials-6+json"
    }

    response = hub.execute_get(components_url, custom_headers=custom_headers)
    oldresponse = hub.execute_get(oldcomponents_url, custom_headers=custom_headers)
    if response.status_code == 200 and oldresponse.status_code == 200:
        components = response.json()
        oldcomponents = oldresponse.json()
        components = components.get("items", [])
        oldcomponents = oldcomponents.get("items", [])

        if args.reviewed or args.unreviewed:
            filter_to = "REVIEWED" if args.reviewed else "NOT_REVIEWED"

            components = list(
                filter(lambda c: c["reviewStatus"] == filter_to, components)
            )
            oldcomponents = list(
                filter(lambda c: c["reviewStatus"] == filter_to, oldcomponents)
            )

        if args.vulnerabilities:
            for component in components:
                vulnerabilities_url = hub.get_link(component, "vulnerabilities")
                response = hub.execute_get(vulnerabilities_url)
                vulnerabilities = []
                if response.status_code == 200:
                    vulnerabilities = response.json().get("items", [])
                component["vulnerabilities"] = vulnerabilities
            for oldcomponent in oldcomponents:
                vulnerabilities_url = hub.get_link(oldcomponent, "vulnerabilities")
                response = hub.execute_get(vulnerabilities_url)
                vulnerabilities = []
                if response.status_code == 200:
                    vulnerabilities = response.json().get("items", [])
                oldcomponent["vulnerabilities"] = vulnerabilities

        if args.custom_fields:
            for component in components:
                custom_fields_url = hub.get_link(component, "custom-fields")
                response = hub.execute_get(
                    custom_fields_url, custom_headers=custom_headers
                )
                custom_fields = []
                if response.status_code == 200:
                    custom_fields = response.json().get("items", [])
                component["custom_fields"] = custom_fields
            for oldcomponent in oldcomponents:
                custom_fields_url = hub.get_link(oldcomponent, "custom-fields")
                response = hub.execute_get(
                    custom_fields_url, custom_headers=custom_headers
                )
                custom_fields = []
                if response.status_code == 200:
                    custom_fields = response.json().get("items", [])
                oldcomponent["custom_fields"] = custom_fields

        componentsNameVersionList = set()
        for c in components:
            # print(c)
            # sys.exit()
            cn = c.get("componentName", "NA")
            cv = c.get("componentVersionName", "NA")
            is_ignored = c.get("ignored", "NA")
            if not is_ignored:
                componentsNameVersionList.add(f"{cn}_{cv}")
        oldcomponentsNameVersionList = set()
        for c in oldcomponents:
            cn = c.get("componentName", "NA")
            cv = c.get("componentVersionName", "NA")
            is_ignored = c.get("ignored", "NA")
            if not is_ignored:
                oldcomponentsNameVersionList.add(f"{cn}_{cv}")

        differenceslist = list(componentsNameVersionList - oldcomponentsNameVersionList)
        from openpyxl import load_workbook
        import pathlib

        this_package = __package__
        source = files(this_package).joinpath("template_aosd_new.xlsx")
        with as_file(source) as excelFile:
            if excelFile:
                exceldir = pathlib.Path(args.o).resolve()
                wb = load_workbook(excelFile)
                n = 0
                row = 3
                for c in components:
                    cn = c.get("componentName", "NA")
                    cv = c.get("componentVersionName", "NA")
                    if f"{cn}_{cv}" in differenceslist:
                        n += 1
                        row += 1
                        licenses = c.get("licenses", [])
                        origins = c.get("origins", [])
                        cl = ""
                        if licenses:
                            for l in licenses:
                                cl = (
                                    f'{l.get("spdxId","NA")}'
                                    if len(licenses) == 1
                                    else f'{cl}, {l.get("spdxId","============")}'
                                )
                        co = ""
                        if origins:
                            for o in origins:
                                co = (
                                    f'{o.get("externalId","============")}'
                                    if len(origins) == 1
                                    else f'{co}, {o.get("externalId","============")}'
                                )
                        print(f"{n} [{co}]  [{cn}] [{cv}] [{cl}]")
                        ws = wb["records"]  # records is the sheet name
                        ws[f"A{row}"] = n
                        ws[f"B{row}"] = co
                        ws[f"C{row}"] = cn
                        ws[f"D{row}"] = cv
                        ws[f"F{row}"] = cl
                        ws[f"H{row}"] = "yes, dynamically"
                        ws[f"I{row}"] = "no"
                        ws[f"J{row}"] = "as object code"
                wb.save(
                    f"{exceldir}/{str(date.today())}_license_overview_only_new.xlsx"
                )
            """
            A3  n
            B3 co
            C3 cn
            D3 cv
            E3 & O3 is url   <--- required 
            F3  cl
            H3 yes, dynamically (linkage)
            I3 no  (modified)
            J3 as object code  (distribution)
            """


if __name__ == "__main__":
    main()
