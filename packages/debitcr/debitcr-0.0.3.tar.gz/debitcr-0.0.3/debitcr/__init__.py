from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
from decimal import Decimal
from time import sleep
import requests
import json


def run():
    OUTPUT_PATH = 'output/'
    
    HEADERS = {
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36" ,
        'referer':'https://www.google.com/'
    }
    
    class Token:
        def __init__(self, symbol, multiplier, api_link='', api_key=''):
            self.symbol = symbol
            self.multiplier = multiplier
            self.api_link = api_link
            self.api_key = api_key

    SCAN_ASSETS = {
        'ETH',
        'AVAX',
        'BNB',
        'MATIC',
        'FTM',
        'USDT',
    }            

    SCAN_APIS = [
        Token('ETH', 1e18, 'etherscan.io', 'XQBQNA7X7Z8BU41N9DZ5WUGT5E7YSFF15M'),
        Token('MATIC', 1e18, 'polygonscan.com', 'FD9PJJDSJ6JX3K2ZKWJIWRJMEWXEBG31QU'),
        Token('BNB', 1e8, 'bscscan.com', 'IG49A7A5VG6SEA2ZBSDEH2NXRHTNDN93B2'),
        Token('AVAX', 1e18, 'snowtrace.io', 'RS48JWVWEP5UYSE79S3XRJACW9DYJCTEDU'),
        Token('FTM', 1e18, 'ftmscan.com', 'ZPCF593WU8S5J7E3VRMKRGBK44NEAEETAU')
    ]

    def call_scan_api(token, address, offset=10000, start_block = 0, end_block = 99999999):
        normal_tx = requests.get(f'https://api.{token.api_link}/api?module=account&action=txlist&address={address}&startblock={start_block}&endblock={end_block}&offset={offset}&sort=desc&apikey={token.api_key}')

        if normal_tx.status_code != 200:
            return False
        
        erc20_transfer = requests.get(f'https://api.{token.api_link}/api?module=account&action=tokentx&address={address}&offset={offset}&startblock={start_block}&endblock={end_block}&sort=desc&apikey={token.api_key}')
        
        responses = [normal_tx, erc20_transfer]

        filename = f'{OUTPUT_PATH}{token.symbol}_{address}.xlsx'
        
        work_book = Workbook()
        work_sheet = work_book.active
        work_sheet.title = "Normal Transactions"
        work_book.create_sheet("ERC20 Transfers")

        for index in range(len(work_book.worksheets)):
            if responses[index].status_code == 200:
                content = json.loads(responses[index].content)
                rows = content['result']
                work_book.active = index
                work_sheet = work_book.active
                work_sheet.append(['Date', 'Block Index', 'Transaction ID', 'Sent', 'Received', 'Asset'])
                
                for row in rows:
                    asset = row['tokenSymbol'] if 'tokenSymbol' in row else token.symbol
                    tokenMultiplier = 10**int(row['tokenDecimal']) if 'tokenDecimal' in row else token.multiplier
                    time = str(datetime.fromtimestamp(int(row['timeStamp'])))
                    sent = int(row['value'])/tokenMultiplier if row['from'].lower() == address.lower() else 0
                    received = int(row['value'])/tokenMultiplier if row['to'].lower() == address.lower() else 0
                    
                    work_sheet.append([time, row['blockNumber'], row['hash'], sent, received, asset])

        work_book.save(filename)
        return True
        
        
    BLOCK_EXPLORER = {
        'BTC' : 'https://bitcoinblockexplorers.com/api/',  
        'BCH' : 'https://bchblockexplorer.com/api/', 
        'LTC' : 'https://litecoinblockexplorer.net/api/', 
        'DASH': 'https://dashblockexplorer.com/api/',
        'DOGE': 'https://dogeblocks.com/api/',
        'ZEC' : 'https://zecblockexplorer.com/api/'
    }
    
    BLOCK_EXPLORER_TXS = 'transactions'
    
    def get_blockexplorer_vin_value(address, vin):
        if address in vin['addresses']:
            return Decimal(vin['value'])
            
        return 0
        
    def get_blockexplorer_vout_value(address, vout):
        if 'scriptPubKey' in vout and 'addresses' in vout['scriptPubKey'] and address in vout['scriptPubKey']['addresses']:
            return Decimal(vout['value'])
        
        return 0    
        
    def call_block_explorer_api(symbol, api_url, address):
        tx_history_url = f'{api_url}address/{address}'
        tx_history = requests.get(tx_history_url, headers=HEADERS)
        tx_history_content = json.loads(tx_history.content)
        
        filename = f'{OUTPUT_PATH}{symbol}_{address}.xlsx'
        
        work_book = Workbook()
        work_sheet = work_book.active
        work_sheet.append(['Date', 'Block Index', 'Transaction ID', 'Sent', 'Received', 'Asset'])
        
        pages = int(tx_history_content['totalPages'])
        for page in range(1, pages+1):
            tx_history_url = f'{api_url}address/{address}?page={page}'
            tx_history = requests.get(tx_history_url, headers=HEADERS)
            
            if tx_history.status_code != 200:
                return False
                
            tx_history_content = json.loads(tx_history.content)
            
            if BLOCK_EXPLORER_TXS not in tx_history_content:
                return False
                
            txs = tx_history_content[BLOCK_EXPLORER_TXS]
            
            for tx in txs:
                tx_url = f'{api_url}tx/{tx.strip()}'
                tx = requests.get(tx_url, headers=HEADERS)
                
                if tx.status_code != 200:
                    return False
                
                tx_content = json.loads(tx.content)
                
                value = 0
                for vin in tx_content['vin']:
                    value -= get_blockexplorer_vin_value(address, vin)

                for vout in tx_content['vout']:              
                    value += get_blockexplorer_vout_value(address, vout)  

                received = value if value > 0 else 0
                sent = value*-1 if value < 0 else 0 
                time = str(datetime.fromtimestamp(int(tx_content['blocktime'])))
                work_sheet.append([time, tx_content['blockheight'], tx_content['txid'], sent, received, symbol])
         
        work_book.save(filename)
        return True
        
    def get_bsv_vin_value(address, vin):
        value = 0
        tx_url = f'https://api.whatsonchain.com/v1/bsv/main/tx/hash/{vin["txid"].strip()}'
        tx = requests.get(tx_url, headers=HEADERS)
        tx_content = json.loads(tx.content)
        
        for vout in tx_content['vout']:
            value += get_blockexplorer_vout_value(address, vout)
            
        return value
        
    def call_bsv_api(symbol, address):
        tx_history = requests.get(f'https://api.whatsonchain.com/v1/bsv/main/address/{address}/history')
        
        if tx_history.status_code !=  200:
            return False
            
        txs = json.loads(tx_history.content)
        
        filename = f'{OUTPUT_PATH}{symbol}_{address}.xlsx'

        work_book = Workbook()
        work_sheet = work_book.active
        work_sheet.append(['Date', 'Block Index', 'Transaction ID', 'Sent', 'Received', 'Asset'])
        
        for tx in txs:
            tx_url = f'https://api.whatsonchain.com/v1/bsv/main/tx/hash/{tx["tx_hash"].strip()}'
            tx = requests.get(tx_url, headers=HEADERS)
            
            if tx.status_code != 200:
                return False            
                
            tx_content = json.loads(tx.content)
 
            time = str(datetime.fromtimestamp(int(tx_content['blocktime'])))
            sent = 0
            received = 0

            for vout in tx_content['vout']:              
                received += get_blockexplorer_vout_value(address, vout)

            if received == 0:
                for vin in tx_content['vin']:
                    sleep(.3)
                    sent += get_bsv_vin_value(address, vin)
                
            work_sheet.append([time, tx_content['blockheight'], tx_content['txid'], sent, received, symbol])        
        work_book.save(filename)    
        return True        
        
    def call_algo_api(token, address):
        txns = requests.get(f'https://algoindexer.algoexplorerapi.io/v2/transactions?address={address}')
        
        if txns.status_code !=  200:
            return False
            
        content = json.loads(txns.content)
        
        if 'message' in content:
            return False

        filename = f'{OUTPUT_PATH}{token.symbol}_{address}.xlsx'
        
        work_book = Workbook()
        work_sheet = work_book.active
        work_sheet.append(['Date', 'Block Index', 'Transaction ID', 'Sent', 'Received', 'Asset'])
        
        rows = content['transactions']
        
        for row in rows:
            time = str(datetime.fromtimestamp(int(row['round-time'])))
            value = int(row['payment-transaction']['amount'])
            sent = value/token.multiplier if row['payment-transaction']['receiver'].lower() != address.lower() else 0
            received = value/token.multiplier if row['payment-transaction']['receiver'].lower() == address.lower() else 0
            
            work_sheet.append([time, row['confirmed-round'], row['id'], sent, received, token.symbol])        
        
        work_book.save(filename)    
        return True
        
    def generate_excel(asset, address):
        print(f'Working with address - {address}')
        print('Generating Excel...')
        success = False
        if asset in SCAN_ASSETS:
            for token in SCAN_APIS:
                if call_scan_api(token, address):
                    success = True
                    break
        elif asset in BLOCK_EXPLORER:
            success = call_block_explorer_api(asset, BLOCK_EXPLORER[asset], address)
        elif asset == 'ALGO':
            success = call_algo_api(Token('ALGO', 1e6), address)
        elif asset == 'BSV':
            success = call_bsv_api(asset, address)
            
        if success:
            print(f'Transactions retrieved successfully as of {datetime.now().strftime("%d/%m/%Y %I:%M:%S %p")}')
        else:
            print('Transactions retrieval failed')
            
    asset = input('Input Asset: ').upper().strip()
    address = input('Input Address: ').strip()
    
    generate_excel(asset, address)


